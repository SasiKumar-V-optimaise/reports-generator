import sqlite3
import pandas as pd
import yaml
import logging
import zipfile
from pathlib import Path
from datetime import datetime, timedelta
from xml.sax.saxutils import escape


logger = logging.getLogger(__name__)


class PipeExporter:
    """
    Shift-wise pipe data exporter (IST timezone).

    Reads:
        config/runtime.yaml

    Outputs:
        outputs/csv/
    """

    IST_OFFSET = ("+5 hours", "+30 minutes")
    DEFAULT_MIN_T_ORIGIN_GAP_SECONDS = 110  # 00:01:50
    DEFAULT_MAX_T_ORIGIN_GAP_SECONDS = 190  # 00:03:10

    def __init__(self):
        self.root = Path(__file__).resolve().parents[2]
        self.cfg = self._load_yaml(self.root / "config" / "runtime.yaml")
        (
            self.min_t_origin_gap_seconds,
            self.max_t_origin_gap_seconds,
        ) = self._diagnosis_t_origin_gap_limits()
        self.min_t_origin_gap_label = self._format_duration(self.min_t_origin_gap_seconds)
        self.max_t_origin_gap_label = self._format_duration(self.max_t_origin_gap_seconds)

        # ---------- DATABASE ----------
        self.db_path = (self.root / self.cfg["database"]["path"]).resolve()
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")

        # ---------- OUTPUT DIR ----------
        csv_dir = self.cfg.get("outputs", {}).get("csv_dir", "outputs/csv")
        self.output_dir = (self.root / csv_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # ---------- SHIFTS ----------
        shifts_cfg = self.cfg.get("history", {}).get("shifts", [])
        if not shifts_cfg:
            raise ValueError("No shifts defined in runtime.yaml")

        self.shifts = {s["name"].lower(): (s["start"], s["end"]) for s in shifts_cfg}

        # Optional: expose last export count
        self.pipe_count: int | None = None

    @staticmethod
    def _load_yaml(path: Path):
        with open(path, "r") as f:
            return yaml.safe_load(f)

    @staticmethod
    def _first_configured(*values):
        for value in values:
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
        return None

    @staticmethod
    def _parse_positive_seconds(value, *, default_seconds: int, name: str) -> int:
        if value is None:
            return default_seconds

        if isinstance(value, (int, float)):
            seconds = int(value)
        else:
            value_text = str(value).strip()
            if not value_text:
                return default_seconds

            if value_text.replace(".", "", 1).isdigit():
                seconds = int(float(value_text))
            else:
                seconds = int(pd.to_timedelta(value_text).total_seconds())

        if seconds <= 0:
            raise ValueError(f"{name} must be greater than 0")

        return seconds

    def _diagnosis_t_origin_gap_limits(self) -> tuple[int, int]:
        diagnosis_cfg = self.cfg.get("diagnosis", {}) or {}

        min_value = self._first_configured(
            diagnosis_cfg.get("t_origin_gap_min"),
            diagnosis_cfg.get("t_origin_gap_min_seconds"),
            diagnosis_cfg.get("t_origin_gap_below"),
            diagnosis_cfg.get("t_origin_gap_below_seconds"),
            self.cfg.get("diagnosis_t_origin_gap_min_seconds"),
        )
        max_value = self._first_configured(
            diagnosis_cfg.get("t_origin_gap_max"),
            diagnosis_cfg.get("t_origin_gap_max_seconds"),
            diagnosis_cfg.get("t_origin_gap_above"),
            diagnosis_cfg.get("t_origin_gap_above_seconds"),
            self.cfg.get("diagnosis_t_origin_gap_max_seconds"),
        )

        min_seconds = self._parse_positive_seconds(
            min_value,
            default_seconds=self.DEFAULT_MIN_T_ORIGIN_GAP_SECONDS,
            name="diagnosis.t_origin_gap_min",
        )
        max_seconds = self._parse_positive_seconds(
            max_value,
            default_seconds=self.DEFAULT_MAX_T_ORIGIN_GAP_SECONDS,
            name="diagnosis.t_origin_gap_max",
        )

        if min_seconds >= max_seconds:
            raise ValueError("diagnosis.t_origin_gap_min must be less than diagnosis.t_origin_gap_max")

        return min_seconds, max_seconds

    def _shift_window(self, date_str: str, shift: str):
        shift = shift.lower()
        if shift not in self.shifts:
            raise ValueError(f"Invalid shift: {shift}")

        start_s, end_s = self.shifts[shift]

        start = datetime.strptime(f"{date_str} {start_s}", "%d-%m-%Y %H:%M")
        end = datetime.strptime(f"{date_str} {end_s}", "%d-%m-%Y %H:%M")

        # Overnight shift (e.g., 22:00 → 06:00)
        if end <= start:
            end += timedelta(days=1)

        return int(start.timestamp()), int(end.timestamp()), start, end

    def _build_query(self):
        h, m = self.IST_OFFSET
        return f"""
        SELECT
            pipe_uid,
            origin,
            pipe_checkpoint,
            datetime(t_origin,'unixepoch','{h}','{m}') AS t_origin,
            datetime(t_loadcell_enter,'unixepoch','{h}','{m}') AS t_loadcell_enter,
            datetime(t_loadcell_exit,'unixepoch','{h}','{m}') AS t_loadcell_exit,
            weight,
            weight_quality,
            weight_samples,
            state,
            datetime(last_seen_ts,'unixepoch','{h}','{m}') AS last_seen_ts
        FROM pipes
        WHERE t_origin BETWEEN ? AND ?
        ORDER BY t_origin DESC;
        """

    def _fetch_shift_df(self, date_str: str, shift: str):
        start_ts, end_ts, start_dt, end_dt = self._shift_window(date_str, shift)

        query = self._build_query()

        with sqlite3.connect(self.db_path) as con:
            df = pd.read_sql_query(query, con, params=(start_ts, end_ts))

        return df, start_dt, end_dt

    @staticmethod
    def _format_duration(seconds):
        if pd.isna(seconds):
            return ""

        total_seconds = int(round(float(seconds)))
        if total_seconds < 0:
            total_seconds = abs(total_seconds)

        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    @staticmethod
    def _missing_mask(series: pd.Series) -> pd.Series:
        as_text = series.astype("string").str.strip().str.lower()
        return series.isna() | as_text.isna() | as_text.isin({"", "none", "nan", "nat", "null"})

    @staticmethod
    def _parse_datetime_series(series: pd.Series) -> pd.Series:
        as_text = series.astype("string").str.strip()
        parsed = pd.to_datetime(as_text, errors="coerce")

        dayfirst_mask = as_text.str.match(r"^\d{2}[-/]\d{2}[-/]\d{4}\b", na=False)
        if dayfirst_mask.any():
            dayfirst_parsed = pd.to_datetime(as_text[dayfirst_mask], errors="coerce", dayfirst=True)
            parsed.loc[dayfirst_mask] = dayfirst_parsed.loc[dayfirst_mask]

        return parsed

    def _build_diagnosis_df(self, df: pd.DataFrame) -> pd.DataFrame:
        diagnosis_df = df.copy()

        if diagnosis_df.empty:
            diagnosis_df["next_pipe_uid"] = pd.Series(dtype="string")
            diagnosis_df["t_origin_gap"] = pd.Series(dtype="string")
            diagnosis_df["t_origin_gap_seconds"] = pd.Series(dtype="float")
            diagnosis_df["t_origin_gap_status"] = pd.Series(dtype="string")
            diagnosis_df["loadcell_status"] = pd.Series(dtype="string")
            diagnosis_df["diagnosis_status"] = pd.Series(dtype="string")
            diagnosis_df["diagnosis_reason"] = pd.Series(dtype="string")
            diagnosis_df["highlight_color"] = pd.Series(dtype="string")
            return diagnosis_df

        origin_dt = self._parse_datetime_series(diagnosis_df["t_origin"])
        next_origin_dt = origin_dt.shift(-1)

        diagnosis_df["next_pipe_uid"] = diagnosis_df["pipe_uid"].shift(-1)
        diagnosis_df["t_origin_gap_seconds"] = (origin_dt - next_origin_dt).dt.total_seconds()
        diagnosis_df["t_origin_gap"] = diagnosis_df["t_origin_gap_seconds"].apply(self._format_duration)

        gap_seconds = diagnosis_df["t_origin_gap_seconds"]
        has_gap = gap_seconds.notna()
        gap_too_fast = has_gap & (gap_seconds < self.min_t_origin_gap_seconds)
        gap_too_slow = has_gap & (gap_seconds > self.max_t_origin_gap_seconds)

        diagnosis_df["t_origin_gap_status"] = "OK"
        diagnosis_df.loc[~has_gap, "t_origin_gap_status"] = "NO_NEXT_PIPE"
        diagnosis_df.loc[gap_too_fast, "t_origin_gap_status"] = "TOO_FAST"
        diagnosis_df.loc[gap_too_slow, "t_origin_gap_status"] = "TOO_SLOW"

        missing_enter = self._missing_mask(diagnosis_df["t_loadcell_enter"])
        missing_exit = self._missing_mask(diagnosis_df["t_loadcell_exit"])
        missing_both = missing_enter & missing_exit

        diagnosis_df["loadcell_status"] = "OK"
        diagnosis_df.loc[missing_enter & ~missing_exit, "loadcell_status"] = "MISSING_ENTRY"
        diagnosis_df.loc[~missing_enter & missing_exit, "loadcell_status"] = "MISSING_EXIT"
        diagnosis_df.loc[missing_both, "loadcell_status"] = "MISSING_ENTRY_AND_EXIT"

        origin_gap_abnormal = gap_too_fast | gap_too_slow
        loadcell_abnormal = missing_enter | missing_exit
        abnormal = origin_gap_abnormal | loadcell_abnormal

        diagnosis_df["diagnosis_status"] = "OK"
        diagnosis_df.loc[abnormal, "diagnosis_status"] = "ABNORMAL"

        reasons = []
        for gap_status, loadcell_status in zip(
            diagnosis_df["t_origin_gap_status"],
            diagnosis_df["loadcell_status"],
        ):
            row_reasons = []
            if gap_status == "TOO_FAST":
                row_reasons.append(f"T_ORIGIN_GAP_BELOW_{self.min_t_origin_gap_label}")
            elif gap_status == "TOO_SLOW":
                row_reasons.append(f"T_ORIGIN_GAP_ABOVE_{self.max_t_origin_gap_label}")

            if loadcell_status == "MISSING_ENTRY":
                row_reasons.append("LOADCELL_ENTRY_MISSING")
            elif loadcell_status == "MISSING_EXIT":
                row_reasons.append("LOADCELL_EXIT_MISSING")
            elif loadcell_status == "MISSING_ENTRY_AND_EXIT":
                row_reasons.append("LOADCELL_ENTRY_AND_EXIT_MISSING")

            reasons.append("; ".join(row_reasons))

        diagnosis_df["diagnosis_reason"] = reasons
        diagnosis_df["highlight_color"] = ""
        diagnosis_df.loc[abnormal, "highlight_color"] = "red"

        return diagnosis_df

    def _write_diagnosis_xlsx(self, diagnosis_df: pd.DataFrame, out_path: Path):
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl is not installed; using built-in XLSX writer")
            self._write_diagnosis_xlsx_basic(diagnosis_df, out_path)
            return

        sheet_name = "Diagnosis"

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            diagnosis_df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            header_fill = PatternFill(fill_type="solid", fgColor="FF1F2937")
            header_font = Font(color="FFFFFFFF", bold=True)
            abnormal_fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
            abnormal_font = Font(color="FFFFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

            abnormal_rows = diagnosis_df.get("diagnosis_status", pd.Series(dtype="string")) == "ABNORMAL"
            for row_idx, is_abnormal in enumerate(abnormal_rows, start=2):
                if not is_abnormal:
                    continue
                for cell in ws[row_idx]:
                    cell.fill = abnormal_fill
                    cell.font = abnormal_font

            for col_idx, column_name in enumerate(diagnosis_df.columns, start=1):
                values = diagnosis_df[column_name].astype("string").fillna("")
                max_value_length = int(values.str.len().max()) if len(values) else 0
                width = min(max(len(str(column_name)), max_value_length) + 2, 45)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    @staticmethod
    def _xlsx_col_name(col_idx: int) -> str:
        name = ""
        while col_idx:
            col_idx, remainder = divmod(col_idx - 1, 26)
            name = chr(65 + remainder) + name
        return name

    @staticmethod
    def _xlsx_cell_xml(row_idx: int, col_idx: int, value, style_id: int = 0) -> str:
        cell_ref = f"{PipeExporter._xlsx_col_name(col_idx)}{row_idx}"
        style = f' s="{style_id}"' if style_id else ""

        if pd.isna(value):
            return f'<c r="{cell_ref}"{style}/>'

        if isinstance(value, bool):
            return f'<c r="{cell_ref}" t="b"{style}><v>{int(value)}</v></c>'

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return f'<c r="{cell_ref}"{style}><v>{value}</v></c>'

        text = escape(str(value))
        return f'<c r="{cell_ref}" t="inlineStr"{style}><is><t>{text}</t></is></c>'

    def _write_diagnosis_xlsx_basic(self, diagnosis_df: pd.DataFrame, out_path: Path):
        columns = [str(c) for c in diagnosis_df.columns]
        max_col = max(len(columns), 1)
        max_row = len(diagnosis_df) + 1
        last_col = self._xlsx_col_name(max_col)
        sheet_ref = f"A1:{last_col}{max_row}"

        widths = []
        for column_name in columns:
            values = diagnosis_df[column_name].astype("string").fillna("")
            max_value_length = int(values.str.len().max()) if len(values) else 0
            widths.append(min(max(len(str(column_name)), max_value_length) + 2, 45))

        cols_xml = "".join(
            f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
            for idx, width in enumerate(widths, start=1)
        )

        header_cells = "".join(
            self._xlsx_cell_xml(1, col_idx, column_name, style_id=1)
            for col_idx, column_name in enumerate(columns, start=1)
        )
        rows_xml = [f'<row r="1">{header_cells}</row>']

        abnormal_rows = diagnosis_df.get("diagnosis_status", pd.Series(dtype="string")) == "ABNORMAL"
        for row_idx, (_, row) in enumerate(diagnosis_df.iterrows(), start=2):
            style_id = 2 if bool(abnormal_rows.iloc[row_idx - 2]) else 0
            cells = "".join(
                self._xlsx_cell_xml(row_idx, col_idx, row[column_name], style_id=style_id)
                for col_idx, column_name in enumerate(diagnosis_df.columns, start=1)
            )
            rows_xml.append(f'<row r="{row_idx}">{cells}</row>')

        worksheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <dimension ref="{sheet_ref}"/>
  <sheetViews>
    <sheetView workbookViewId="0">
      <pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>
    </sheetView>
  </sheetViews>
  <cols>{cols_xml}</cols>
  <sheetData>{''.join(rows_xml)}</sheetData>
  <autoFilter ref="{sheet_ref}"/>
  <pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>
</worksheet>'''

        workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Diagnosis" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>'''

        workbook_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''

        root_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''

        content_types_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>'''

        styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3">
    <font><sz val="11"/><name val="Calibri"/></font>
    <font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font>
    <font><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font>
  </fonts>
  <fills count="4">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF1F2937"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFF0000"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1"/>
    <xf numFmtId="0" fontId="2" fillId="3" borderId="0" xfId="0" applyFont="1" applyFill="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
  <dxfs count="0"/>
  <tableStyles count="0" defaultTableStyle="TableStyleMedium9" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>'''

        with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as xlsx:
            xlsx.writestr("[Content_Types].xml", content_types_xml)
            xlsx.writestr("_rels/.rels", root_rels_xml)
            xlsx.writestr("xl/workbook.xml", workbook_xml)
            xlsx.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
            xlsx.writestr("xl/styles.xml", styles_xml)
            xlsx.writestr("xl/worksheets/sheet1.xml", worksheet_xml)

    def export_diagnosis(self, date_str: str, shift: str):
        logger.info("Fetching pipe diagnosis data (date=%s, shift=%s)...", date_str, shift)

        df, start_dt, end_dt = self._fetch_shift_df(date_str, shift)
        diagnosis_df = self._build_diagnosis_df(df)

        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"pipes_diagnosis_{date_str.replace('-','')}_{shift.lower()}_{timestamp}.xlsx"
        out_path = self.output_dir / filename
        self._write_diagnosis_xlsx(diagnosis_df, out_path)

        gap_status = diagnosis_df.get("t_origin_gap_status", pd.Series(dtype="string"))
        loadcell_status = diagnosis_df.get("loadcell_status", pd.Series(dtype="string"))
        diagnosis_status = diagnosis_df.get("diagnosis_status", pd.Series(dtype="string"))

        summary = {
            "pipe_count": int(len(diagnosis_df)),
            "abnormal_count": int((diagnosis_status == "ABNORMAL").sum()),
            "t_origin_gap_abnormal_count": int(gap_status.isin(["TOO_FAST", "TOO_SLOW"]).sum()),
            "t_origin_gap_too_slow_count": int((gap_status == "TOO_SLOW").sum()),
            "t_origin_gap_too_fast_count": int((gap_status == "TOO_FAST").sum()),
            "t_origin_gap_min_seconds": int(self.min_t_origin_gap_seconds),
            "t_origin_gap_max_seconds": int(self.max_t_origin_gap_seconds),
            "t_origin_gap_min_label": self.min_t_origin_gap_label,
            "t_origin_gap_max_label": self.max_t_origin_gap_label,
            "loadcell_missing_count": int((loadcell_status != "OK").sum()),
        }

        logger.info(
            "PIPE DIAGNOSIS | shift=%s | from=%s | to=%s | abnormal=%s | saved=%s",
            shift.upper(), start_dt, end_dt, summary["abnormal_count"], out_path
        )
        return out_path, summary

    def export(self, date_str: str, shift: str):
        logger.info("Fetching pipe data (date=%s, shift=%s)...", date_str, shift)

        df, start_dt, end_dt = self._fetch_shift_df(date_str, shift)

        # ---------- FILE NAME ----------
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"pipes_{date_str.replace('-','')}_{shift.lower()}_{timestamp}.csv"
        out_path = self.output_dir / filename
        df.to_csv(out_path, index=False)

        pipe_count = int(len(df))
        self.pipe_count = pipe_count

        logger.info(
            "PIPE REPORT | shift=%s | from=%s | to=%s | pipe_count=%s | saved=%s",
            shift.upper(), start_dt, end_dt, pipe_count, out_path
        )
        return out_path, pipe_count
